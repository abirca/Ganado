# Librerías estándar de Python
import calendar
import io
import json
import locale
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

# Librerías de terceros
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Librerías de Django
from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import Http404, get_object_or_404, redirect, render
from django.utils.safestring import mark_safe

# Imports locales
from .forms import GastoForm, MovimientoClienteForm, MovimientoForm, ProveedorForm


RUTA_EXCEL = settings.RUTA_EXCEL
RUTA_EXCEL_SEGUNDO =  settings.RUTA_EXCEL_SEGUNDO
# Configurar a español (Colombia por ejemplo)
locale.setlocale(locale.LC_TIME, 'es_CO.UTF-8') 
# Configuración para tipos de entidad (Proveedores/Clientes)
ENTITY_CONFIG = {
    'proveedor': {
        'form': MovimientoForm,
        'movimiento_form_template': 'formulario.html',
        'resumen_template': 'resumen.html',
        'movimientos_template': 'movimientos.html',
        'agregar_persona_template': 'agregar_persona.html',
        'dashboard_template': 'dashboard.html',
        'editar_template': 'editar.html',
        'sheet_movimientos': 'Proveedores',
        'sheet_resumen': 'Resumen',
        'url_index': 'mi_app:movimiento_proveedor_agregar',
        'url_resumen': 'mi_app:proveedores_resumen',
    },
    'cliente': {
        'form': MovimientoClienteForm,
        'movimiento_form_template': 'formularioCliente.html',
        'resumen_template': 'resumenCliente.html',
        'movimientos_template': 'movimientosCliente.html',
        'agregar_persona_template': 'agregar_personaCliente.html',
        'dashboard_template': 'dashboardCliente.html',
        'editar_template': 'editarCliente.html',
        'sheet_movimientos': 'ProveedoresCliente',
        'sheet_resumen': 'ResumenCliente',
        'url_index': 'mi_app:movimiento_cliente_agregar',
        'url_resumen': 'mi_app:clientes_resumen',
    },
    'gastos': {
        'sheet_gastos': 'Gastos',
        'url_index': 'mi_app:gastos_lista',
        'form': GastoForm,
    }
}

# Funciones genéricas reutilizables
def cargar_datos_excel(sheet_name):
    """Carga datos desde una hoja de Excel específica"""
    if not os.path.exists(RUTA_EXCEL):
        return []
    
    try:
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        if sheet_name not in wb.sheetnames:
            return []
        
        ws = wb[sheet_name]
        return list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        print(f"Error al cargar datos de Excel: {e}")
        return []

def obtener_ultimo_id(sheet_name):
    """Obtiene el último ID utilizado en una hoja de Excel"""
    datos = cargar_datos_excel(sheet_name)
    if not datos:
        return 0
    
    # Buscar el máximo ID en la primera columna
    max_id = 0
    for fila in datos:
        if fila and len(fila) > 0 and isinstance(fila[0], (int, float)):
            max_id = max(max_id, int(fila[0]))
    
    return max_id

def generar_id_factura(proveedor, sheet_name):
    """Genera un IdFactura incremental tipo F-001, F-002, ..."""
    movimientos = cargar_datos_excel(sheet_name)
    max_num = 0
    for mov in movimientos:
        if len(mov) >= 7 and mov[2] == proveedor:  # Asegurarse de que hay al menos 7 columnas y que IdFactura no es None
            id_factura = mov[6]  # columna IdFactura
            if id_factura and isinstance(id_factura, str) and id_factura.startswith('F-'):
                try:
                    num = int(id_factura.split('-')[1])
                    max_num = max(max_num, num)
                except:
                    continue
    nuevo_id = f"F-{max_num + 1:03d}"
    return nuevo_id

def guardar_en_excel(sheet_name, datos, encabezados=None, modo='overwrite'):
    """Guarda datos en una hoja de Excel específica"""
    try:
        if os.path.exists(RUTA_EXCEL):
            wb = load_workbook(RUTA_EXCEL)
        else:
            wb = openpyxl.Workbook()
            # Eliminar la hoja por defecto si existe
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            if encabezados:
                ws.append(encabezados)
        else:
            ws = wb[sheet_name]
            # Si el modo es 'overwrite', limpiar la hoja
            if modo == 'overwrite':
                # Eliminar todas las filas excepto los encabezados
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
                # Si no hay encabezados, mantener los existentes
                if encabezados:
                    # Reemplazar encabezados existentes
                    for idx, encabezado in enumerate(encabezados, 1):
                        ws.cell(row=1, column=idx, value=encabezado)
        
        # Agregar datos
        for dato in datos:
            ws.append(dato)
        
        wb.save(RUTA_EXCEL)
        wb.save(RUTA_EXCEL_SEGUNDO)
        return True
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")
        return False

def normalizar_total(total_raw):
    """Normaliza el valor total eliminando caracteres no numéricos"""
    try:
        total_str = re.sub(r'[^\d]', '', total_raw) if total_raw else '0'
        return Decimal(total_str) if total_str else Decimal('0')
    except Exception as e:
        print(f"Error al normalizar total: {e}")
        return Decimal('0')

def normalizar_fecha(fecha_input):
    if isinstance(fecha_input, str):
        # Convertir cadena a datetime.date
        return datetime.strptime(fecha_input, '%Y-%m-%d').date()
    elif isinstance(fecha_input, datetime):
        # Si ya es datetime, extraer solo la fecha
        return fecha_input.date()
    elif isinstance(fecha_input, date):
        # Si ya es date, devolverlo tal cual
        return fecha_input
    else:
        raise TypeError("El tipo de entrada debe ser str, datetime o date")

def obtener_movimientos_filtrados(entity_type, proveedor_filtrado=None, fecha_filtrada=None, estado_filtrado=None):
    """Obtiene movimientos filtrados por proveedor, fecha y estado (activo/inactivo)"""
    config = ENTITY_CONFIG[entity_type]
    movimientos = []
    
    datos = cargar_datos_excel(config['sheet_movimientos'])
    
    for row in datos:
        # Asegurarse de que la fila tiene suficientes columnas
        if len(row) < 8:
            estado = None # No tiene columna de estado
        else:
            estado = row[7].lower()

        # Mapear los datos de la fila a un diccionario
        mov = {
            'id': row[0],
            'fecha': row[1],
            'proveedor': row[2],
            'detalle': row[3],
            'obs': row[4],
            'total': row[5],
            'idfactura': row[6] if len(row) > 6 else '',
            'estado': estado
        }
        
        # Aplicar filtros
        cumple_proveedor = (not proveedor_filtrado or (mov['proveedor'] and str(mov['proveedor']).strip().lower() == proveedor_filtrado.strip().lower()))
        cumple_fecha = (not fecha_filtrada or normalizar_fecha(mov['fecha']) == normalizar_fecha(fecha_filtrada))
        cumple_estado = (not estado_filtrado or (mov['estado'] and mov['estado'] == estado_filtrado.lower()))
        
        if cumple_proveedor and cumple_fecha and cumple_estado:
            movimientos.append(mov)
    
    return movimientos

def obtener_resumen_filtrado(entity_type, proveedor_filtrado=None):
    """Obtiene resumen filtrado por proveedor"""
    config = ENTITY_CONFIG[entity_type]
    datos = []
    
    resumen_data = cargar_datos_excel(config['sheet_resumen'])
    
    for row in resumen_data:
        if len(row) < 5:
            continue
        
        id = row[0]
        proveedor = row[1]
        factura = row[2] or 0
        ahorro = row[3] or 0
        saldo = row[4] or 0
        
        if proveedor_filtrado:
            if proveedor == proveedor_filtrado:
                datos.append((id, proveedor, factura, ahorro, saldo))
        else:
            datos.append((id, proveedor, factura, ahorro, saldo))
    
    return datos

def recalcular_resumen(entity_type):
    """Recalcula el resumen considerando solo facturas y abonos activos"""
    config = ENTITY_CONFIG[entity_type]

    movimientos = [list(row) for row in cargar_datos_excel(config['sheet_movimientos'])]

    # Filtrar solo los movimientos activos
    movimientos_activos = [mov for mov in movimientos if len(mov) >= 8 and mov[7].lower() == 'activa']

    resumen_dict = {}

    for mov in movimientos_activos:
        _, fecha, proveedor, detalle, obs, total, id_factura, estado = mov

        if proveedor not in resumen_dict:
            resumen_dict[proveedor] = {
                'facturas': Decimal('0'),
                'abonos': Decimal('0'),
                'saldo': Decimal('0')
            }

        if detalle.lower() == 'factura':
            resumen_dict[proveedor]['facturas'] += Decimal(str(total))
        elif detalle.lower() == 'abono':
            resumen_dict[proveedor]['abonos'] += Decimal(str(total))

        # Calcular saldo actual
        resumen_dict[proveedor]['saldo'] = resumen_dict[proveedor]['facturas'] - resumen_dict[proveedor]['abonos']

    # Preparar datos para Excel
    resumen_data = []
    for idx, (proveedor, valores) in enumerate(resumen_dict.items(), start=1):
        resumen_data.append([
            idx,  # ID incremental
            proveedor,
            float(valores['facturas']),
            float(valores['abonos']),
            float(valores['saldo'])
        ])

    encabezados = ['Id', 'Proveedor', 'Total Facturas', 'Total Abonos', 'Saldo']
    guardar_en_excel(config['sheet_resumen'], resumen_data, encabezados, modo='overwrite')

    return True

def calcular_saldo_factura(id_factura, movimientos_data):
    """Calcular el saldo pendiente de una factura específica"""
    saldo = 0
    
    for movimiento in movimientos_data:
        if len(movimiento) > 6 and movimiento[6] == id_factura:
            if 'factura' in movimiento[3].lower():
                saldo += float(movimiento[5])
                print=(f"Saldo inicial de la factura {id_factura}: {saldo}")
            elif 'abono' in movimiento[3].lower():
                saldo -= float(movimiento[5])
                print =(f"Abono aplicado a la factura {id_factura}: {movimiento[5]}, saldo restante: {saldo}")

    return saldo  # No permitir saldos negativos

def actualizar_estado_facturas(proveedor, id_factura, sheet_name):
    """Actualizar el estado de las facturas de un proveedor"""
    movimientos_data = [list(row) for row in cargar_datos_excel(sheet_name)]

    for i, movimiento in enumerate(movimientos_data):
        # Check for the correct provider, invoice detail, and matching invoice ID
        if len(movimiento) > 7 and \
            movimiento[2] == proveedor and \
            movimiento[6] == id_factura:
            
            # Check the current status and update only if it's 'Activa'
            if movimiento[7].lower() == 'activa':
                movimientos_data[i][7] = 'Inactiva'

    return movimientos_data

# Vistas para Gastos
def gastos(request):
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    gastos_list = []
    
    for gasto in gastos_data:
        if len(gasto) >= 6:  # Asegurarse de que tiene todos los campos
            gastos_list.append({
                'id': gasto[0],
                'fecha': gasto[1],
                'categoria': gasto[2],
                'placa': gasto[3],
                'conductor': gasto[4],
                'precio': gasto[5],
            })
    
    # Filtros
    categoria_filtro = request.GET.get('categoria', '')
    placa_filtro = request.GET.get('placa', '')
    fecha_filtro = request.GET.get('fecha', '')
    
    if categoria_filtro:
        gastos_list = [g for g in gastos_list if g['categoria'] == categoria_filtro]
    if placa_filtro:
        gastos_list = [g for g in gastos_list if placa_filtro.lower() in g['placa'].lower()]
    if fecha_filtro:
        gastos_list = [g for g in gastos_list if normalizar_fecha(g['fecha']) == normalizar_fecha(fecha_filtro)]
    
    # Paginación
    paginator = Paginator(gastos_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'categoria_filtro': categoria_filtro,
        'placa_filtro': placa_filtro,
        'fecha_filtro': fecha_filtro,
    }
    return render(request, 'gastos.html', context)

def agregar_gasto(request):
    # Cargar gastos existentes
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    if request.method == 'POST':
        form = GastoForm(request.POST)
        if form.is_valid():
            # Obtener datos del formulario
            fecha = form.cleaned_data['fecha']
            categoria = form.cleaned_data['categoria']
            placa = form.cleaned_data['placa'] or ''
            conductor = form.cleaned_data['conductor'] or ''
            precio = form.cleaned_data['precio']
            
            
            # Obtener último ID
            nuevo_id = obtener_ultimo_id(config['sheet_gastos']) + 1
            fecha = normalizar_fecha(fecha)
            
            # Nueva fila
            nueva_fila = [nuevo_id, fecha, categoria, placa, conductor, float(precio)]
            
            # Agregar a los datos
            gastos_data.append(nueva_fila)
            
            # Guardar en Excel
            encabezados = ['Id', 'Fecha', 'Categoria', 'Placa', 'Conductor', 'Precio']

            if guardar_en_excel(config['sheet_gastos'],gastos_data, encabezados, modo='overwrite'):
                messages.success(request, 'Gasto agregado correctamente.')
                return redirect('mi_app:gasto_agregar')  # Redirigir a la misma página para ver el resumen actualizado
            else:
                messages.error(request, 'Error al guardar el gasto. Inténtelo de nuevo.')
    else:
        form = GastoForm()
        
    # Convertir a objetos Gasto para facilitar el procesamiento
    gastos = []
    for gasto in gastos_data:
        if len(gasto) >= 6:  # Asegurarse de que hay suficientes columnas
            gastos.append({
                'fecha': normalizar_fecha(gasto[1]),
                'categoria': gasto[2],
                'placa': gasto[3],
                'conductor': gasto[4],
                'precio': float(gasto[5]) if gasto[5] else 0
            })
    
    # Calcular totales
    hoy = date.today()
    total_hoy = sum(g['precio'] for g in gastos if g['fecha'] == hoy)
    total_mes = sum(g['precio'] for g in gastos if g['fecha'].month == hoy.month and g['fecha'].year == hoy.year)
    
    # Resumen por categoría
    categorias = {}
    for gasto in gastos:
        if gasto['categoria'] not in categorias:
            categorias[gasto['categoria']] = 0
        categorias[gasto['categoria']] += gasto['precio']
    
    total_general = sum(categorias.values())
    resumen_categorias = []
    for categoria, total in categorias.items():
        porcentaje = (total / total_general * 100) if total_general > 0 else 0
        resumen_categorias.append({
            'categoria': categoria,
            'total': total,
            'porcentaje': porcentaje
        })
    
    # Últimos 5 gastos
    ultimos_gastos = sorted(gastos, key=lambda x: x['fecha'], reverse=True)[:5]
    
    context = {
        'form': form,
        'total_hoy': total_hoy,
        'total_mes': total_mes,
        'categorias_count': len(categorias),
        'total_gastos': len(gastos),
        'resumen_categorias': resumen_categorias,
        'ultimos_gastos': ultimos_gastos
    }
    
    return render(request, 'agregar_gasto.html', context)

def editar_gasto(request, id):
    # Cargar gastos
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    gasto_editar = None
    
    for gasto in gastos_data:
        if gasto[0] == id:
            gasto_editar = gasto
            break
    
    if not gasto_editar:
        messages.error(request, 'Gasto no encontrado.')
        return redirect('gastos')
    
    # Convertir la fecha al formato correcto para el formulario
    
    fecha_valor = gasto_editar[1].strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        form = GastoForm(request.POST)
        if form.is_valid():
            # Actualizar los datos del gasto
            fecha = form.cleaned_data['fecha']
            categoria = form.cleaned_data['categoria']
            placa = form.cleaned_data['placa'] or ''
            conductor = form.cleaned_data['conductor'] or ''
            precio = form.cleaned_data['precio']
            
            fecha = normalizar_fecha(fecha)

            # Actualizar la fila en gastos_data
            for i, gasto in enumerate(gastos_data):
                if gasto[0] == id:
                    gastos_data[i] = [id, fecha, categoria, placa, conductor, float(precio)]
                    break
            
            # Guardar en Excel
            encabezados = ['Id', 'Fecha', 'Categoria', 'Placa', 'Conductor', 'Precio']
            if guardar_en_excel(config['sheet_gastos'], gastos_data, encabezados, modo='overwrite'):
                messages.success(request, 'Gasto actualizado correctamente.')
                return redirect('gastos')
            else:
                messages.error(request, 'Error al actualizar el gasto.')
    else:
        # Rellenar el formulario con los datos del gasto
        form = GastoForm(initial={
            'fecha': fecha_valor,
            'categoria': gasto_editar[2],
            'placa': gasto_editar[3],
            'conductor': gasto_editar[4],
            'precio': gasto_editar[5],
        })
    
    return render(request, 'editar_gasto.html', {'form': form, 'id': id})

def eliminar_gasto(request, id):
    # Cargar gastos existentes
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    
    # Buscar el gasto por ID
    gasto_encontrado = None
    for gasto in gastos_data:
        if gasto[0] == id:
            gasto_encontrado = gasto
            break
    
    if gasto_encontrado is None:
        messages.error(request, 'El gasto no existe.')
        return redirect('gastos')
    
    if request.method == 'POST':
        # Eliminar el gasto de la lista
        gastos_data.remove(gasto_encontrado)
        
        # Guardar los datos actualizados
        encabezados = ['Id', 'Fecha', 'Categoria', 'Placa', 'Conductor', 'Precio']
        if guardar_en_excel(config['sheet_gastos'], gastos_data, encabezados, modo='overwrite'):
            messages.success(request, 'Gasto eliminado correctamente.')
        else:
            messages.error(request, 'Error al eliminar el gasto.')
        
        return redirect('gastos')
    
    # Convertir a diccionario para mostrar en la plantilla
    gasto_dict = {
        'id': gasto_encontrado[0],
        'fecha': gasto_encontrado[1].date() if isinstance(gasto_encontrado[1], datetime) else gasto_encontrado[1],
        'categoria': gasto_encontrado[2],
        'placa': gasto_encontrado[3],
        'conductor': gasto_encontrado[4],
        'precio': gasto_encontrado[5]
    }
    
    return render(request, 'eliminar_gasto.html', {'gasto': gasto_dict})

def resumen_gastos(request):
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    
    # Verificar si se solicita descargar Excel
    download = request.GET.get('download', '')
    
    # Convertir a lista de diccionarios
    gastos_list = []
    for gasto in gastos_data:
        if len(gasto) >= 6:
            gastos_list.append({
                'id': gasto[0],
                'fecha': normalizar_fecha(gasto[1]),
                'categoria': gasto[2],
                'placa': gasto[3],
                'conductor': gasto[4],
                'precio': float(gasto[5]) if gasto[5] else 0,
            })
    
    # Filtros
    categoria_filtro = request.GET.get('categoria', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    if categoria_filtro:
        gastos_list = [g for g in gastos_list if g['categoria'] == categoria_filtro]
        
    if fecha_inicio:
        gastos_list = [g for g in gastos_list if g['fecha'] >= normalizar_fecha(fecha_inicio)]

    if fecha_fin:
        gastos_list = [ g for g in gastos_list if g['fecha'] <= normalizar_fecha(fecha_fin)]

    # Ordenar por fecha ascendente
    gastos_list.sort(key=lambda x: x['fecha'])
    
    # Agrupar por categoría
    resumen_categoria = {}
    for gasto in gastos_list:
        cat = gasto['categoria']
        resumen_categoria[cat] = resumen_categoria.get(cat, 0) + gasto['precio']
    
    # Calcular el total general
    total_general = sum(resumen_categoria.values())
    
    # Convertir a lista para la plantilla, agregando el porcentaje
    resumen = []
    for k, v in resumen_categoria.items():
        porcentaje = (v / total_general * 100) if total_general != 0 else 0
        resumen.append({
            'categoria': k, 
            'total': v,
            'porcentaje': porcentaje
        })
    
    # Si se solicita descargar Excel
    if download == 'excel':
        return generar_excel_gastos(gastos_list, resumen_categoria)
    
    # Gastos por mes
    gastos_por_mes = {}
    for gasto in gastos_list:       
        mes_key = f"{gasto['fecha'].year}-{gasto['fecha'].month:02d}"
        if mes_key not in gastos_por_mes:
            gastos_por_mes[mes_key] = 0
        gastos_por_mes[mes_key] += gasto['precio']
    
    # Formatear gastos por mes para la plantilla
    gastos_por_mes_formateados = []
    for mes_key, total in gastos_por_mes.items():
        año, mes = mes_key.split('-')
        nombre_mes = datetime.strptime(mes, '%m').strftime('%B')
        gastos_por_mes_formateados.append({
            'ano': año,
            'mes': nombre_mes,
            'total': total
        })
    
    context = {
        'resumen': resumen,
        'gastos_por_mes': gastos_por_mes_formateados,
        'categoria_filtro': categoria_filtro,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_general': total_general  # Usamos el total_general que ya calculamos
    }
    return render(request, 'resumen_gastos.html', context)

def generar_excel_gastos(gastos_list, resumen_categoria):
    """Genera un archivo Excel con los gastos y totales por categoría con estilos aplicados"""
    # Crear DataFrame con los gastos
    df_gastos = pd.DataFrame(gastos_list)
    
    # Eliminar columna ID si existe
    if 'id' in df_gastos.columns:
        df_gastos = df_gastos.drop(columns=['id'])

    
    # Crear DataFrame con los totales por categoría
    df_totales = pd.DataFrame([
        {'categoria': 'TOTAL ' + cat.upper(), 'precio': total, 'fecha': '', 'placa': '', 'conductor': ''}
        for cat, total in resumen_categoria.items()
    ])
    
    # Agregar fila de total general
    total_general = sum(resumen_categoria.values())
    df_total_general = pd.DataFrame([{
        'categoria': 'TOTAL GENERAL',
        'precio': total_general,
        'fecha': '',
        'placa': '',
        'conductor': ''
    }])
    
    # Concatenar todos los DataFrames
    df_final = pd.concat([df_gastos, df_totales, df_total_general], ignore_index=True)
    
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Gastos"
    
    # Definir estilos según la imagen
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul más oscuro
    data_font = Font(size=11)
    total_font = Font(bold=True, color="000000", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Azul claro
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_right = Alignment(horizontal="right", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    
    # Escribir el título
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="RESUMEN GASTOS")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir datos en la hoja (empezando desde la fila 3)
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Aplicar estilos a la fila de encabezado (fila 3)
            if r_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
            else:
                # Aplicar estilos a filas de datos
                cell.font = data_font
                
                # Aplicar estilos a filas de totales
                if str(value).startswith('TOTAL'):
                    cell.font = total_font
                    cell.fill = total_fill
                    # Para las celdas de total, alinear categoría a la izquierda
                    if c_idx == 2:  # Columna de categoría
                        cell.alignment = alignment_left
                    elif c_idx == 5:  # Columna de precio
                        cell.alignment = alignment_right
                        if value != 'precio':  # No aplicar a la celda de encabezado
                            # Formato de moneda chilena
                            cell.number_format = '"$"#,##0.00'
                    else:
                        cell.alignment = alignment_center
                else:
                    # Para datos normales
                    if c_idx == 1:  # Columna de fecha
                        cell.alignment = alignment_center
                    elif c_idx == 2:  # Columna de categoría
                        cell.alignment = alignment_left
                    elif c_idx == 5:  # Columna de precio
                        cell.alignment = alignment_right
                        # Formato de moneda chilena
                        cell.number_format = '"$"#,##0.00'
                    else:
                        cell.alignment = alignment_center
                    
            # Aplicar bordes a todas las celdas
            cell.border = border
    
    # Ajustar anchos de columnas según la imagen
    column_widths = {
        'A': 15,  # fecha
        'B': 15,  # categoria
        'C': 12,  # placa
        'D': 15,  # conductor
        'E': 15   # precio
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resumen_gastos.xlsx"'
    
    # Guardar el libro de trabajo en la respuesta
    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
    
    return response

def dashboard_gastos(request):
    config = ENTITY_CONFIG['gastos']
    gastos_data = cargar_datos_excel(config['sheet_gastos'])
    gastos_list = []
    
    for gasto in gastos_data:
        if len(gasto) >= 6:
            gastos_list.append({
                'fecha': normalizar_fecha(gasto[1]),
                'categoria': gasto[2],
                'placa': gasto[3],
                'conductor': gasto[4],
                'precio': gasto[5],
            })
    
    # Aplicar filtros
    categoria_filtro = request.GET.get('categoria', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    placa_filtro = request.GET.get('placa', '')
    
    
    gastos_filtrados = gastos_list
    
    if categoria_filtro:
        gastos_filtrados = [g for g in gastos_filtrados if g['categoria'] == categoria_filtro]
    
    if fecha_inicio:
        gastos_filtrados = [g for g in gastos_filtrados if g['fecha'] >= normalizar_fecha(fecha_inicio)]
    
    if fecha_fin:
        gastos_filtrados = [g for g in gastos_filtrados if g['fecha'] <= normalizar_fecha(fecha_fin)]
    
    if placa_filtro:
        gastos_filtrados = [g for g in gastos_filtrados if g['placa'] == placa_filtro]
    
    # Gastos por categoría
    gastos_por_categoria = defaultdict(float)
    for gasto in gastos_filtrados:
        cat = gasto['categoria']
        gastos_por_categoria[cat] += gasto['precio']
    
    # Top 5 gastos más altos
    top_gastos = sorted(gastos_filtrados, key=lambda x: x['precio'], reverse=True)[:5]
    
    # Gastos por mes
    gastos_por_mes = defaultdict(float)
    for gasto in gastos_filtrados:
        try:
            fecha = gasto['fecha']
            mes_key = f"{fecha.year}-{fecha.month:02d}"
            gastos_por_mes[mes_key] += gasto['precio']
        except:
            continue
    
    # Preparar datos para gráficos
    meses_ordenados = sorted(gastos_por_mes.keys())
    meses_labels = [f"{calendar.month_abbr[int(m.split('-')[1])]} {m.split('-')[0]}" for m in meses_ordenados]
    gastos_mensuales = [gastos_por_mes[m] for m in meses_ordenados]
    
    # Gastos por placa
    gastos_por_placa = defaultdict(float)
    for gasto in gastos_filtrados:
        if gasto['placa']:
            gastos_por_placa[gasto['placa']] += gasto['precio']
    
    # Top 5 placas con más gastos
    top_placas = sorted(gastos_por_placa.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Gastos por conductor
    gastos_por_conductor = defaultdict(float)
    for gasto in gastos_filtrados:
        if gasto['conductor']:
            gastos_por_conductor[gasto['conductor']] += gasto['precio']
    
    # Top 5 conductores con más gastos
    top_conductores = sorted(gastos_por_conductor.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Gastos por categoría y mes (para gráfico de barras apiladas)
    gastos_por_categoria_mes = defaultdict(lambda: defaultdict(float))
    categorias = set()
    
    for gasto in gastos_filtrados:
        try:
            fecha = gasto['fecha']
            mes_key = f"{fecha.year}-{fecha.month:02d}"
            categoria = gasto['categoria']
            categorias.add(categoria)
            gastos_por_categoria_mes[mes_key][categoria] += gasto['precio']
        except:
            continue
    
    # Preparar datos para gráfico de barras apiladas
    datasets_apiladas = []
    colores = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796']
    
    for i, categoria in enumerate(sorted(categorias)):
        color = colores[i % len(colores)]
        data = []
        for mes in meses_ordenados:
            data.append(gastos_por_categoria_mes[mes].get(categoria, 0))
        
        datasets_apiladas.append({
            'label': categoria,
            'data': data,
            'backgroundColor': color,
        })
    
    # Calcular totales y promedios
    total_gastos = sum([g['precio'] for g in gastos_filtrados])
    
    # Promedio mensual
    num_meses = len(meses_ordenados) or 1
    promedio_mensual = total_gastos / num_meses
    
    # Categoría con mayor gasto
    categoria_mayor_gasto = max(gastos_por_categoria.items(), key=lambda x: x[1], default=('N/A', 0))
    
    # Lista de placas únicas para el filtro
    placas_unicas = sorted(set([g['placa'] for g in gastos_list if g['placa']]))
    
    context = {
        'gastos_por_categoria': gastos_por_categoria,
        'top_gastos': top_gastos,
        'total_gastos': total_gastos,
        'promedio_mensual': promedio_mensual,
        'categoria_mayor_gasto': categoria_mayor_gasto[0],
        'categorias': list(gastos_por_categoria.keys()),
        'totales_por_categoria': list(gastos_por_categoria.values()),
        'meses': meses_labels,
        'gastos_por_mes': gastos_mensuales,
        'top_placas': top_placas,
        'top_conductores': top_conductores,
        'datasets_apiladas': datasets_apiladas,
        'meses_apiladas': meses_labels,
        'categoria_filtro': categoria_filtro,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'placa_filtro': placa_filtro,
        'placas_unicas': placas_unicas,
    }
    
    return render(request, 'dashboard_gastos.html', context)

# Vistas genéricas
def movimiento_view(request, entity_type):
    """Vista genérica para movimientos de proveedores o clientes"""
    config = ENTITY_CONFIG[entity_type]
    form = config['form']()
    
    proveedor_filtrado = request.GET.get('proveedor', None)
    fecha_filtrada = request.GET.get('fecha', None)
    
    movimientos = obtener_movimientos_filtrados(entity_type, proveedor_filtrado, fecha_filtrada,"Activa")
    resumen = cargar_datos_excel(config['sheet_resumen'])
    resumen_filtrado = obtener_resumen_filtrado(entity_type, proveedor_filtrado)
    
    movimientos.sort(key=lambda x: x['fecha'] if x['fecha'] else datetime.min.date(), reverse=True)

    paginator = Paginator(movimientos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, config['movimiento_form_template'], {
        'form': form,
        'resumen': resumen,
        'movimientos': page_obj,
        'resumen_Filtrado': resumen_filtrado,
        'proveedor_filtrado': proveedor_filtrado,
        'fecha_filtrada': fecha_filtrada,
        'paginator': paginator,
        'page_obj': page_obj,
    })

def resumen_view(request, entity_type):
    """Vista genérica para resumen de proveedores o clientes"""
    config = ENTITY_CONFIG[entity_type]
    
    proveedor_filtrado = request.GET.get('proveedor', None)
    datos = obtener_resumen_filtrado(entity_type, proveedor_filtrado)
    
    return render(request, config['resumen_template'], {
        'resumen': datos,
        'proveedor_filtrado': proveedor_filtrado,
    })

def movimientos_list_view(request, entity_type):
    """Vista genérica para listar movimientos de proveedores o clientes con filtros mejorados"""
    config = ENTITY_CONFIG[entity_type]
    
    # Obtener todos los parámetros de filtro
    proveedor_filtrado = request.GET.get('proveedor', '').strip()
    estado_filtrado = request.GET.get('estado', '').strip()
    id_factura_filtrado = request.GET.get('id_factura', '').strip()
    fecha_filtrada = request.GET.get('fecha', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    
    movimientos = []
    datos = cargar_datos_excel(config['sheet_movimientos'])
    resumen = cargar_datos_excel(config['sheet_resumen'])
    
    # Obtener lista única de proveedores para el filtro
    proveedores_unicos = set()
    for row in resumen:
        if len(row) > 1:  # Asegurarse de que hay al menos 2 columnas
            proveedores_unicos.add(row[1])
    
    for mov_data in datos:
        # Saltar filas que no tienen suficientes columnas
        if len(mov_data) < 6:  # Mínimo necesario para los campos básicos
            continue
        
        # Asegurar que tenemos al menos 8 columnas (rellenar con valores vacíos si es necesario)
        while len(mov_data) < 8:
            mov_data.append('')
            
        mov_id, fecha_raw, proveedor, detalle, obs, total, id_factura, estado = mov_data
        
        
        # Filtrar por proveedor
        if proveedor_filtrado and (not proveedor or str(proveedor).strip().lower() != proveedor_filtrado.lower()):
            continue
            
        # Filtrar por estado
        if estado_filtrado and (not estado or str(estado).strip().lower() != estado_filtrado.lower()):
            continue
            
        # Filtrar por ID de factura
        if id_factura_filtrado and (not id_factura or str(id_factura).strip().lower() != id_factura_filtrado.lower()):
            continue

        # Si hay fecha específica, ignorar rango
        if fecha_filtrada:
            if not fecha_raw or normalizar_fecha(fecha_raw) != normalizar_fecha(fecha_filtrada):
                continue
        else:
            # Filtrar por rango de fechas
            if fecha_inicio and (not fecha_raw or normalizar_fecha(fecha_raw) < normalizar_fecha(fecha_inicio)):
                continue

            if fecha_fin and (not fecha_raw or normalizar_fecha(fecha_raw) > normalizar_fecha(fecha_fin)):
                continue
        
        # Si llegamos aquí, el movimiento pasa todos los filtros
        movimientos.append({
            'id': mov_id,
            'fecha': fecha_raw.date() if fecha_raw else None,
            'proveedor': proveedor,
            'detalle': detalle,
            'obs': obs,
            'total': total,
            'id_factura': id_factura,
            'estado': estado
        })
    
    # Ordenar movimientos por fecha (más reciente primero)
    movimientos.sort(key=lambda x: x['fecha'] if x['fecha'] else datetime.min.date(), reverse=True)
    
    # Paginación
    paginator = Paginator(movimientos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preparar parámetros para mantener en la paginación
    all_params = request.GET.copy()
    if 'page' in all_params:
        del all_params['page']
    
    return render(request, config['movimientos_template'], {
        'resumen': resumen,
        'proveedores_unicos': sorted(proveedores_unicos),
        'movimientos': page_obj,
        'proveedor_filtrado': proveedor_filtrado,
        'estado_filtrado': estado_filtrado,
        'id_factura_filtrado': id_factura_filtrado,
        'fecha_filtrada': fecha_filtrada,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'paginator': paginator,
        'page_obj': page_obj,
        'all_params': all_params.urlencode()  # Para mantener todos los parámetros en los enlaces de paginación
    })

def agregar_persona_view(request, entity_type):
    """Vista genérica para agregar personas (proveedores o clientes)"""
    config = ENTITY_CONFIG[entity_type]
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre'].strip()
            
            # Validar si ya existe en Excel
            resumen_data = cargar_datos_excel(config['sheet_resumen'])
            
            existe = False
            for row in resumen_data:
                if len(row) > 1 and row[1] == nombre:
                    existe = True
                    break
            
            if existe:
                messages.info(request, "La persona ya existe en el archivo Excel. ❌")
                form.add_error('nombre', 'La persona ya existe en el archivo Excel.')
            else:
                # Obtener el último ID y generar uno nuevo
                nuevo_id = obtener_ultimo_id(config['sheet_resumen']) + 1
                
                # Crear nueva fila para el resumen
                nueva_fila = [nuevo_id, nombre, 0, 0, 0]
                
                # Agregar a los datos existentes
                resumen_data.append(nueva_fila)
                
                # Guardar en Excel
                encabezados = ['Id', 'Proveedor', 'Total Facturas', 'Total Abonos', 'Saldo']
                if guardar_en_excel(config['sheet_resumen'], resumen_data, encabezados, modo='overwrite'):
                    messages.success(request, f'Persona {nombre} agregada correctamente. ✔️')
                    return redirect(config['url_resumen'])
                else:
                    messages.error(request, 'Error al guardar la persona. Inténtelo de nuevo. ❌')
                
                return redirect(config['url_resumen'])
    else:
        form = ProveedorForm()
    
    return render(request, config['agregar_persona_template'], {'form': form})

def dashboard_view(request, entity_type):
    """Vista genérica para dashboard de proveedores o clientes"""
    config = ENTITY_CONFIG[entity_type]
    
    # Obtener parámetros de filtro
    proveedor_filtrado = request.GET.get('proveedor', None)
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    id_factura_filtrado = request.GET.get('id_factura', None)  # Nuevo filtro
    
    
    proveedores = []
    facturas_totales = []
    abonos_totales = []
    saldos_totales = []
    
    facturas_por_mes = {}
    abonos_por_mes = {}
    
    # Nuevos datos para las gráficas KPI
    total_facturado = 0
    total_abonado = 0
    porcentaje_abono = 0
    
    # Cargar datos de resumen
    resumen_data = cargar_datos_excel(config['sheet_resumen'])
    for row in resumen_data:
        if len(row) < 5:
            continue
            
        id, proveedor, factura, abonos, saldo = row[0], row[1], row[2] or 0, row[3] or 0, row[4] or 0
        
        # Filtrar por proveedor si se especificó
        if proveedor_filtrado and proveedor != proveedor_filtrado:
            continue
            
        proveedores.append(proveedor)
        facturas_totales.append(factura)
        abonos_totales.append(abonos)
        saldos_totales.append(saldo)
        
        # Calcular totales para KPI
        total_facturado += factura
        total_abonado += abonos
    
    # Calcular porcentaje de abono
    if total_facturado > 0:
        porcentaje_abono = (total_abonado / total_facturado) * 100
    
    # Cargar datos de movimientos
    movimientos_data = cargar_datos_excel(config['sheet_movimientos'])
    for row in movimientos_data:
        if len(row) < 8:  # Asegurarse de que hay al menos 8 columnas (incluyendo id_factura)
            continue
            
        id_, fecha_str, prov, detalle, obs, total, id_factura, _ = row
        
        # Filtrar por ID de factura si se especificó
        if id_factura_filtrado and id_factura != id_factura_filtrado:
            continue
            
        # Filtrar por proveedor si se especificó
        if proveedor_filtrado and prov != proveedor_filtrado:
            continue
        
        try:
            fecha = normalizar_fecha(fecha_str)
        except Exception:
            continue  # si no se puede convertir, se salta

        # Filtrar por rango de fechas si se especificó
        if fecha_inicio_str and fecha < normalizar_fecha(fecha_inicio_str):
            continue
        if fecha_fin_str and fecha > normalizar_fecha(fecha_fin_str):
            continue
        
        # Convertir fecha a mes-año
        mes_ano = fecha.strftime('%Y-%m')
        
        # Facturas (detalle == 'Factura')
        if detalle == 'Factura':
            facturas_por_mes.setdefault(prov, {})
            facturas_por_mes[prov][mes_ano] = facturas_por_mes[prov].get(mes_ano, 0) + (total or 0)
        
        # Abonos (detalle == 'Abono')
        if detalle == 'Abono':
            abonos_por_mes.setdefault(prov, {})
            abonos_por_mes[prov][mes_ano] = abonos_por_mes[prov].get(mes_ano, 0) + (total or 0)
    
    # Meses únicos combinando facturas y abonos
    meses = set()
    for prov_data in facturas_por_mes.values():
        meses.update(prov_data.keys())
    for prov_data in abonos_por_mes.values():
        meses.update(prov_data.keys())
    meses = sorted(meses)
    
    # Formatear datos para Chart.js
    facturas_linea = []
    abonos_linea = []
    proveedores_filtrados = sorted(set(list(facturas_por_mes.keys()) + list(abonos_por_mes.keys())))
    
    for prov in proveedores_filtrados:
        # Facturas
        data_fact = [facturas_por_mes.get(prov, {}).get(mes, 0) for mes in meses]
        facturas_linea.append({'proveedor': prov, 'datos': data_fact})
        # Abonos
        data_ahorro = [abonos_por_mes.get(prov, {}).get(mes, 0) for mes in meses]
        abonos_linea.append({'proveedor': prov, 'datos': data_ahorro})
    
    # Calcular totales por mes para la evolución temporal
    facturas_por_mes_totales = {mes: 0 for mes in meses}
    abonos_por_mes_totales = {mes: 0 for mes in meses}
    
    for prov in facturas_por_mes:
        for mes, valor in facturas_por_mes[prov].items():
            facturas_por_mes_totales[mes] += valor
    
    for prov in abonos_por_mes:
        for mes, valor in abonos_por_mes[prov].items():
            abonos_por_mes_totales[mes] += valor
    
    context = {
        'proveedores': proveedores,
        'facturas': mark_safe(json.dumps(facturas_totales)),
        'abonos': mark_safe(json.dumps(abonos_totales)),
        'saldos': mark_safe(json.dumps(saldos_totales)),
        'proveedores_filtrados': proveedores_filtrados,
        'meses': mark_safe(json.dumps(meses)),
        'facturas_linea': mark_safe(json.dumps(facturas_linea)),
        'abonos_linea': mark_safe(json.dumps(abonos_linea)),
        'proveedor_filtrado': proveedor_filtrado or '',
        'id_factura_filtrado': id_factura_filtrado or '',  # Nuevo valor para el formulario
        # Nuevos datos para las gráficas
        'total_facturado': total_facturado,
        'total_abonado': total_abonado,
        'porcentaje_abono': porcentaje_abono,
        'facturas_por_mes_totales': mark_safe(json.dumps([facturas_por_mes_totales[mes] for mes in meses])),
        'abonos_por_mes_totales': mark_safe(json.dumps([abonos_por_mes_totales[mes] for mes in meses])),
        # Valores actuales de filtros para mostrarlos en el formulario
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
    }
    return render(request, config['dashboard_template'], context)

def guardar_movimiento_view(request, entity_type):
    config = ENTITY_CONFIG[entity_type]

    if request.method == 'POST':
        form = config['form'](request.POST)
        if form.is_valid():
            data = form.cleaned_data
            proveedor = data['proveedor']
            detalle = data['detalle']
            obs = data['obs']
            fecha = normalizar_fecha(data['fecha'])
            total = normalizar_total(request.POST.get('total', ''))

            movimientos_data = cargar_datos_excel(config['sheet_movimientos'])
            
            if detalle.lower() == 'factura':
                filtra_data = [mov for mov in movimientos_data if len(mov) >= 8 and mov[2] == proveedor and mov[7].lower() == 'activa']
                saldo_anterior = 0

                if filtra_data:
                    # If there is an active invoice, get the first (and only) one.
                    factura_activa = filtra_data[0]
                    id_anterior = factura_activa[6]
                    
                    # Calculate the previous balance using the filtered data.
                    saldo_anterior = calcular_saldo_factura(id_anterior, filtra_data)
                    
                    # Update the status of the previous invoice.
                    movimientos_data = actualizar_estado_facturas(proveedor, id_anterior, config['sheet_movimientos'])

                if saldo_anterior != 0:
                    total_mensaje = f"Total {obs} : {int(total)}"
                    total +=  Decimal(str(saldo_anterior))
                    obs = f"{total_mensaje} + saldo anterior {int(saldo_anterior)} = total {int(total)}"
                    if total < 0:
                        messages.error(request, f"No se puede guardar la factura, debe ser mayor al saldo: {int(saldo_anterior)}.")
                        return redirect(config['url_index'])

                id_factura = generar_id_factura(proveedor, config['sheet_movimientos'])
                # Nueva fila de abono
                id_fila = obtener_ultimo_id(config['sheet_movimientos']) + 1

                nueva_fila = [
                    id_fila,
                    fecha,
                    proveedor,
                    detalle,
                    obs,
                    float(total),
                    id_factura,
                    'Activa'  # La factura sigue activa
                ]

                movimientos_data.append(nueva_fila)

            elif detalle.lower() == 'abono':
                # Buscar factura activa del proveedor
                id_factura_activa = None
                for row in movimientos_data:
                    if len(row) >= 8 and row[2] == proveedor and row[7] == 'Activa' and row[3].lower() == 'factura':
                        id_factura_activa = row[6]
                        break

                if not id_factura_activa:
                    messages.error(request, 'No existe una factura activa para aplicar el abono.')
                    return redirect(config['url_index'])

                # Nueva fila de abono
                id_fila = obtener_ultimo_id(config['sheet_movimientos']) + 1
                nueva_fila = [
                    id_fila,
                    fecha,
                    proveedor,
                    detalle,
                    obs,
                    float(total),
                    id_factura_activa,
                    'Activa'  # La factura sigue activa
                ]

                movimientos_data.append(nueva_fila)

            encabezados = ['Id', 'Fecha', 'Proveedor', 'Detalle', 'Obs', 'Total', 'IdFactura', 'Estado']
            if guardar_en_excel(config['sheet_movimientos'], movimientos_data, encabezados, modo='overwrite'):
                recalcular_resumen(entity_type)
                messages.success(request, f'Movimiento de {detalle} para {proveedor} guardado correctamente.')
                return redirect(config['url_index'])
            else:
                messages.error(request, 'Error al guardar el movimiento. Inténtelo de nuevo.')

    else:
        form = config['form']()
        messages.info(request, 'Por favor, complete el formulario para agregar un nuevo movimiento.')

    return render(request, config['movimiento_form_template'], {'form': form})

def editar_movimiento_view(request, entity_type, index):
    """Vista genérica para editar movimientos de proveedores o clientes"""
    config = ENTITY_CONFIG[entity_type]
    
    # Buscar el movimiento en Excel
    movimientos_data = cargar_datos_excel(config['sheet_movimientos'])
    mov = None
    
    for row in movimientos_data:
        if row and len(row) > 0 and row[0] == index:
            mov = {
                'id': row[0],
                'fecha': row[1],
                'proveedor': row[2],
                'detalle': row[3],
                'obs': row[4],
                'total': row[5],
                'idfactura': row[6],
                'estado': row[7]
            }
            break
    
    if not mov:
        messages.error(request, 'Movimiento no encontrado.')
        raise Http404("Movimiento no encontrado")
    
    if request.method == 'POST':
        form = config['form'](request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            # Actualizar el movimiento en la lista
            for i, row in enumerate(movimientos_data):
                if row and len(row) > 0 and row[0] == index:
                    movimientos_data[i] = [
                        index,
                        normalizar_fecha(data['fecha']),
                        data['proveedor'],
                        data['detalle'],
                        data['obs'],
                        float(data['total']),
                        mov['idfactura'],  # Preserve the original 'IdFactura'
                        mov['estado'] 
                    ]
                    break
            
            # Guardar en Excel
            encabezados = ['Id', 'Fecha', 'Proveedor', 'Detalle', 'Obs', 'Total', 'IdFactura', 'Estado']
            if guardar_en_excel(config['sheet_movimientos'], movimientos_data, encabezados, modo='overwrite'):
            
                # Recalcular el resumen
                recalcular_resumen(entity_type)
                # Mensaje de éxito
                messages.success(request, f'Se movimiento Modifico el registro correctamente.')

                return redirect(config['url_index'])
            else:
                 # Mensaje de error
                messages.error(request, 'Error al guardar el movimiento. Inténtelo de nuevo.')
    
    else:
        # Convertir fecha a formato string para el formulario
        if isinstance(mov['fecha'], datetime):
            fecha_str = mov['fecha'].strftime('%Y-%m-%d')
            
        initial_data = {
            'fecha': fecha_str,
            'proveedor': mov['proveedor'],
            'detalle': mov['detalle'],
            'obs': mov['obs'],
            'total': mov['total'],
        }
        form = config['form'](initial=initial_data)

        
    
    return render(request, config['editar_template'], {'form': form, 'id': index})

def descargar_excel_entidad(request, entity_type):
    """Vista para descargar un archivo Excel con los movimientos de una entidad específica con filtros completos"""
    try:
        config = ENTITY_CONFIG[entity_type]
        
        # Obtener todos los parámetros de filtro
        nombre_entidad = request.GET.get('proveedor', '').strip()
        estado_filtrado = request.GET.get('estado', '').strip()
        id_factura_filtrado = request.GET.get('id_factura', '').strip()
        fecha_especifica = request.GET.get('fecha', '').strip()
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()
        
        # Obtener todos los movimientos y luego filtrar
        movimientos_data = cargar_datos_excel(config['sheet_movimientos'])
        movimientos_filtrados = []
        
        for row in movimientos_data:
            # Asegurar que tenemos al menos 8 columnas
            if len(row) < 6:
                continue
                
            # Rellenar con valores vacíos si es necesario
            while len(row) < 8:
                row.append('')
                
            mov_id, fecha_raw, proveedor, detalle, obs, total, id_factura, estado = row
            
            mov = {
                'id': mov_id,
                'fecha': fecha_raw,
                'proveedor': proveedor,
                'detalle': detalle,
                'obs': obs,
                'total': total,
                'id_factura': id_factura,
                'estado': estado
            }
            
            # Aplicar filtros
            cumple_proveedor = True
            cumple_estado = True
            cumple_id_factura = True
            cumple_fecha_especifica = True
            cumple_fecha_inicio = True
            cumple_fecha_fin = True
            
            # Filtro por proveedor
            if nombre_entidad:
                prov = mov['proveedor'] or ''
                cumple_proveedor = (str(prov).strip().lower() == nombre_entidad.strip().lower())
            
            # Filtro por estado
            if estado_filtrado:
                est = mov['estado'] or ''
                cumple_estado = (str(est).strip().lower() == estado_filtrado.strip().lower())
            
            # Filtro por ID de factura
            if id_factura_filtrado:
                id_fact = mov['id_factura'] or ''
                cumple_id_factura = (str(id_fact).strip().lower() == id_factura_filtrado.strip().lower())
            
            # Normalizar la fecha del movimiento
            try:
                fecha_mov = normalizar_fecha(mov['fecha'])
            except Exception:
                fecha_mov = None

            # Inicializar banderas
            cumple_fecha_especifica = True
            cumple_fecha_inicio = True
            cumple_fecha_fin = True

            # Filtro por fecha específica
            if fecha_especifica:
                try:
                    fecha_especifica_dt = normalizar_fecha(fecha_especifica)
                    cumple_fecha_especifica = fecha_mov and (fecha_mov == fecha_especifica_dt)
                except Exception:
                    cumple_fecha_especifica = False
            else:
                # Filtro por rango de fechas
                if fecha_inicio:
                    try:
                        fecha_inicio_dt = normalizar_fecha(fecha_inicio)
                        cumple_fecha_inicio = fecha_mov and (fecha_mov >= fecha_inicio_dt)
                    except Exception:
                        cumple_fecha_inicio = False

                if fecha_fin:
                    try:
                        fecha_fin_dt = normalizar_fecha(fecha_fin)
                        cumple_fecha_fin = fecha_mov and (fecha_mov <= fecha_fin_dt)
                    except Exception:
                        cumple_fecha_fin = False

            # Determinar qué filtros aplicar
            if fecha_especifica:
                # Solo usar filtro de fecha específica
                if cumple_proveedor and cumple_estado and cumple_id_factura and cumple_fecha_especifica:
                    movimientos_filtrados.append(mov)
            else:
                # Usar filtros de rango de fechas
                if cumple_proveedor and cumple_estado and cumple_id_factura and cumple_fecha_inicio and cumple_fecha_fin:
                    movimientos_filtrados.append(mov)

        
        movimientos_filtrados.sort(key=lambda x: (x['id_factura'] or '', x['fecha']))

        # Agrupar movimientos por ID de factura para calcular saldos
        movimientos_por_factura = {}
        for mov in movimientos_filtrados:
            id_factura = mov['id_factura'] or 'SIN_FACTURA'
            if id_factura not in movimientos_por_factura:
                movimientos_por_factura[id_factura] = []
            movimientos_por_factura[id_factura].append(mov)
        
        # Calcular saldos por factura
        saldos_por_factura = {}
        for id_factura, movimientos in movimientos_por_factura.items():
            saldo = 0
            for mov in movimientos:
                total_valor = float(mov['total'] or 0)
                if 'factura' in str(mov['detalle']).lower():
                    saldo -= total_valor
                elif 'abono' in str(mov['detalle']).lower():
                    saldo += total_valor
            saldos_por_factura[id_factura] = saldo
        
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Crear un estilo para formato de moneda
        moneda_style = NamedStyle(name="moneda_style")
        moneda_style.number_format = '"$"#,##0.00'
        
        # Si el estilo ya existe, eliminarlo para evitar conflictos
        if "moneda_style" in wb.named_styles:
            del wb.named_styles[wb.named_styles.index("moneda_style")]
        
        # Agregar el estilo al libro
        wb.add_named_style(moneda_style)
        
        # Crear estilos para saldos - CORREGIDOS
        saldo_positivo_style = NamedStyle(name="saldo_positivo_style")
        saldo_positivo_style.number_format = '"$"#,##0.00'
        saldo_positivo_style.font = Font(bold=True, color="007500")  # Verde para saldo a favor
        saldo_positivo_style.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        saldo_negativo_style = NamedStyle(name="saldo_negativo_style")
        saldo_negativo_style.number_format = '"$"#,##0.00'
        saldo_negativo_style.font = Font(bold=True, color="FF0000")  # Rojo para saldo en contra
        saldo_negativo_style.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Agregar estilos de saldo al libro
        for style in [saldo_positivo_style, saldo_negativo_style]:
            if style.name in wb.named_styles:
                del wb.named_styles[wb.named_styles.index(style.name)]
            wb.add_named_style(style)
        
        # Título basado en los filtros aplicados
        titulo = f"Movimientos de {entity_type.capitalize()}"
        if nombre_entidad:
            titulo += f" - {nombre_entidad}"
        if estado_filtrado:
            titulo += f" - Estado: {estado_filtrado}"
        if id_factura_filtrado:
            titulo += f" - ID Factura: {id_factura_filtrado}"
        if fecha_especifica:
            titulo += f" - Fecha: {fecha_especifica}"
        elif fecha_inicio or fecha_fin:
            titulo += " - Período: "
            if fecha_inicio and fecha_fin:
                titulo += f"{fecha_inicio} al {fecha_fin}"
            elif fecha_inicio:
                titulo += f"{fecha_inicio} en adelante"
            elif fecha_fin:
                titulo += f"hasta {fecha_fin}"
        
        ws.title = titulo[:31]  # Limitar a 31 caracteres (máximo de Excel)
        
        # Agregar título
        ws.cell(row=1, column=1, value=titulo)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        
        # Definir encabezados (incluyendo los nuevos campos)
        encabezados = ['Id', 'Fecha', 'Proveedor/Cliente', 'Detalle', 'Observaciones', 'Total', 'ID Factura', 'Estado']
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=3, column=col_num, value=encabezado)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal='center')
        
        # Llenar datos
        fila = 4
        total_facturas = 0
        total_abonos = 0
        id_factura_anterior = None
        
        # Ordenar movimientos por ID de factura para agruparlos
        movimientos_filtrados.sort(key=lambda x: x['id_factura'] or '')
        
        for mov in movimientos_filtrados:
            id_factura_actual = mov['id_factura'] or 'SIN_FACTURA'
            
            # Si cambió el ID de factura, agregar fila de saldo
            if id_factura_anterior and id_factura_anterior != id_factura_actual:
                # Agregar fila de saldo para la factura anterior
                saldo = saldos_por_factura.get(id_factura_anterior, 0)
                ws.cell(row=fila, column=4, value=f"SALDO FACTURA {id_factura_anterior}:")
                ws.cell(row=fila, column=4).font = Font(bold=True)
                
                celda_saldo = ws.cell(row=fila, column=6, value=saldo)
                # CORRECCIÓN: Saldo positivo (a favor) -> verde, Saldo negativo (en contra) -> rojo
                if saldo < 0:  # Saldo negativo (en contra)
                    celda_saldo.style = saldo_negativo_style
                else:  # Saldo positivo (a favor)
                    celda_saldo.style = saldo_positivo_style
                
                # Aplicar fondo gris a toda la fila
                for col in range(1, 9):
                    ws.cell(row=fila, column=col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                fila += 1
            
            # ID
            ws.cell(row=fila, column=1, value=mov['id'])
            
            # Fecha
            fecha = mov['fecha']
            if isinstance(fecha, datetime):
                fecha = fecha.strftime('%d/%m/%Y')
            elif isinstance(fecha, str):
                # Intentar convertir string a fecha y luego formatear
                try:
                    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
                    fecha = fecha_dt.strftime('%d/%m/%Y')
                except:
                    try:
                        fecha_dt = datetime.strptime(fecha, "%d/%m/%Y")
                        fecha = fecha_dt.strftime('%d/%m/%Y')
                    except:
                        try:
                            fecha_dt = datetime.strptime(fecha, "%d-%m-%Y")
                            fecha = fecha_dt.strftime('%d/%m/%Y')
                        except:
                            fecha = "Fecha inválida"
            ws.cell(row=fila, column=2, value=fecha)
            
            # Proveedor/Cliente
            ws.cell(row=fila, column=3, value=mov['proveedor'])
            
            # Detalle
            ws.cell(row=fila, column=4, value=mov['detalle'])
            
            # Observaciones
            ws.cell(row=fila, column=5, value=mov['obs'] or '')
            
            # Total - aplicar formato de moneda
            total_valor = float(mov['total'] or 0)
            celda_total = ws.cell(row=fila, column=6, value=total_valor)
            celda_total.style = moneda_style
            
            # ID Factura
            ws.cell(row=fila, column=7, value=mov['id_factura'] or '')
            
            # Estado
            ws.cell(row=fila, column=8, value=mov['estado'] or '')
            
            # Sumar a totales - CORREGIDO: Usar siempre el mismo valor para consistencia
            if 'factura' in str(mov['detalle']).lower():
                obs = str(mov.get('obs', '')).lower()
                match = re.search(r"(?:total\s+factura|factura)\s*:\s*([\d.,]+)", obs, re.IGNORECASE)
                if match:
                    # Convert the extracted string to float, handling commas as decimals
                    total_valor = float(match.group(1).replace(",", "."))

                total_facturas += total_valor
            elif 'abono' in str(mov['detalle']).lower():
                total_abonos += total_valor
            
            fila += 1
            id_factura_anterior = id_factura_actual
        
        # Agregar saldo para la última factura
        if id_factura_anterior:
            saldo = saldos_por_factura.get(id_factura_anterior, 0)
            ws.cell(row=fila, column=4, value=f"SALDO FACTURA {id_factura_anterior}:")
            ws.cell(row=fila, column=4).font = Font(bold=True)
            
            celda_saldo = ws.cell(row=fila, column=6, value=saldo)
            # CORRECCIÓN: Saldo positivo (a favor) -> verde, Saldo negativo (en contra) -> rojo
            if saldo < 0:  # Saldo negativo (en contra)
                celda_saldo.style = saldo_negativo_style
            else:  # Saldo positivo (a favor)
                celda_saldo.style = saldo_positivo_style
            
            # Aplicar fondo gris a toda la fila
            for col in range(1, 9):
                ws.cell(row=fila, column=col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            fila += 1
        
        # Agregar fila de totales
        fila_total = fila + 1
        
        # Total Facturas
        ws.cell(row=fila_total, column=4, value="TOTAL FACTURAS:")
        ws.cell(row=fila_total, column=4).font = Font(bold=True)
        celda_total_facturas = ws.cell(row=fila_total, column=6, value=total_facturas)
        celda_total_facturas.font = Font(bold=True)
        celda_total_facturas.style = moneda_style
        
        # Total Abonos
        ws.cell(row=fila_total+1, column=4, value="TOTAL ABONOS:")
        ws.cell(row=fila_total+1, column=4).font = Font(bold=True)
        celda_total_abonos = ws.cell(row=fila_total+1, column=6, value=total_abonos)
        celda_total_abonos.font = Font(bold=True)
        celda_total_abonos.style = moneda_style
        
        # Saldo (Diferencia) - CORREGIDO: Facturas menos abonos
        saldo = total_abonos - total_facturas
        ws.cell(row=fila_total+2, column=4, value="SALDO PENDIENTE:")
        ws.cell(row=fila_total+2, column=4).font = Font(bold=True)
        celda_saldo = ws.cell(row=fila_total+2, column=6, value=saldo)
        
        # CORRECCIÓN: Saldo positivo (a favor) -> verde, Saldo negativo (en contra) -> rojo
        if saldo < 0:  # Saldo negativo (en contra)
            celda_saldo.style = saldo_negativo_style
        else:  # Saldo positivo (a favor)
            celda_saldo.style = saldo_positivo_style
            
        
        # Ajustar el ancho de las columnas
        column_widths = [8, 12, 20, 15, 30, 15, 15, 12]  # Anchos personalizados para cada columna
        for column, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(column)
            ws.column_dimensions[col_letter].width = width
        
        # Preparar la respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"movimientos_{entity_type}_{nombre_entidad or 'todos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Guardar el libro en la respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        # En caso de error, retornar una respuesta de error
        import traceback
        error_msg = f"Error al generar el archivo Excel: {str(e)}\n{traceback.format_exc()}"
        return HttpResponse(error_msg, status=500)

def editar_persona_view(request, entity_type, id=None):
    """Vista genérica para agregar o editar personas (proveedores o clientes) buscando en Excel"""
    config = ENTITY_CONFIG[entity_type]
    
    # Si se proporciona un ID, estamos editando una persona existente
    if id:
        # Buscar la persona en el archivo Excel
        persona = None
        es_edicion = True
        nombre_original = None
        
        # Cargar datos de Excel
        resumen_data = cargar_datos_excel(config['sheet_resumen'])
        
        # Buscar la persona por ID
        for row in resumen_data:
            if len(row) > 0 and row[0] == id:
                persona = {
                    'id': row[0],
                    'proveedor': row[1],
                    'facturas': row[2] if len(row) > 2 else 0,
                    'abonos': row[3] if len(row) > 3 else 0,
                    'saldo': row[4] if len(row) > 4 else 0
                }
                nombre_original = row[1]  # Guardar el nombre original para luego
                break
                
        if not persona:
            # Si no se encuentra en Excel, mostrar error 404
            messages.error(request, "La persona no existe en el archivo Excel.")
            raise Http404("La persona no existe")
    else:
        persona = None
        es_edicion = False
        nombre_original = None
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            nombre_nuevo = form.cleaned_data['nombre'].strip()
            
            # Validar si ya existe en Excel (excepto si estamos editando la misma persona)
            resumen_data = cargar_datos_excel(config['sheet_resumen'])
            existe_excel = False
            
            for row in resumen_data:
                if len(row) > 1 and row[1] == nombre_nuevo:  # Nombre está en la segunda columna
                    if es_edicion and row[0] == id:
                        # Es la misma persona que estamos editando, no es un duplicado
                        continue
                    existe_excel = True
                    break
            
            if existe_excel:
                messages.info(request, "La persona ya existe en el archivo Excel. ❌")
                form.add_error('nombre', 'La persona ya existe en el archivo Excel.')
            else:
                # Actualizar el archivo Excel - HOJA DE RESUMEN
                resumen_data = cargar_datos_excel(config['sheet_resumen'])
                
                if es_edicion:
                    # Actualizar la fila existente en los datos de Excel
                    for i, row in enumerate(resumen_data):
                        if len(row) > 0 and row[0] == id:
                            resumen_data[i] = [
                                id,
                                nombre_nuevo,
                                persona['facturas'],
                                persona['abonos'],
                                persona['saldo']
                            ]
                            break
                else:
                    # Crear nueva persona
                    nuevo_id = obtener_ultimo_id(config['sheet_resumen']) + 1
                    nueva_fila = [nuevo_id, nombre_nuevo, 0, 0, 0]
                    resumen_data.append(nueva_fila)
                
                # Guardar los datos actualizados en Excel - HOJA DE RESUMEN
                encabezados = ['Id', 'Proveedor', 'Total Facturas', 'Total Abonos', 'Saldo']
                if guardar_en_excel(config['sheet_resumen'], resumen_data, encabezados, modo='overwrite'):
                    
                    # Si estamos editando, también actualizar la hoja de movimientos
                    if es_edicion and nombre_original != nombre_nuevo:
                        # Cargar datos de movimientos
                        movimientos_data = cargar_datos_excel(config['sheet_movimientos'])
                        # Convert to list of lists if necessary
                        movimientos_data = [list(row) for row in movimientos_data]
                        # Actualizar todos los registros con el nombre antiguo
                        for i, row in enumerate(movimientos_data):
                            if len(row) > 2 and row[2] == nombre_original:  # Nombre en columna 2
                                movimientos_data[i][2] = nombre_nuevo  # Actualizar nombre
                        
                        # Guardar los datos actualizados en Excel - HOJA DE MOVIMIENTOS
                        encabezados_mov = ['Id', 'Fecha', 'Proveedor', 'Detalle', 'Obs', 'Total', 'Id_Factura', 'Estado']
                        if not guardar_en_excel(config['sheet_movimientos'], movimientos_data, encabezados_mov, modo='overwrite'):
                            messages.warning(request, f'Persona actualizada en resumen pero hubo un error al actualizar los movimientos. ❌')
                            return redirect(config['url_resumen'])
                    
                    messages.success(request, f'Persona {nombre_nuevo} {"actualizada" if es_edicion else "agregada"} correctamente. ✔️')
                    return redirect(config['url_resumen'])
                else:
                    messages.error(request, 'Error al guardar la persona. Inténtelo de nuevo. ❌')
    else:
        # Si es edición, precargar el formulario con los datos existentes
        if es_edicion:
            form = ProveedorForm(initial={'nombre': persona['proveedor']})
        else:
            form = ProveedorForm()
    
    return render(request, config['agregar_persona_template'], {
        'form': form,
        'es_edicion': es_edicion,
        'persona': persona
    })

def index_view(request, entity_type_proveedor, entity_type_cliente, entity_type_gastos):
    """Vista para la página principal con estadísticas generales"""
    # Cargar datos de proveedores
    config_proveedor = ENTITY_CONFIG[entity_type_proveedor]
    config_cliente = ENTITY_CONFIG[entity_type_cliente]
    config_gastos = ENTITY_CONFIG[entity_type_gastos]

    resumen_proveedores = cargar_datos_excel(config_proveedor['sheet_resumen'])
    movimientos_proveedores = cargar_datos_excel(config_proveedor['sheet_movimientos'])
    
    # Cargar datos de clientes (si existen)
    try:
        resumen_clientes = cargar_datos_excel(config_cliente['sheet_resumen'])
        movimientos_clientes = cargar_datos_excel(config_cliente['sheet_movimientos'])
    except:
        resumen_clientes = []
        movimientos_clientes = []
    
    # Cargar datos de gastos
    try:
        movimientos_gastos = cargar_datos_excel(config_gastos['sheet_gastos'])
    except:
        movimientos_gastos = []
    
    # Calcular estadísticas generales de proveedores
    total_facturado_proveedores = 0
    total_abonado_proveedores = 0
    saldo_total_proveedores = 0
    proveedores_con_saldo = []
    
    for row in resumen_proveedores:
        if len(row) < 5:
            continue
        factura = row[2] or 0
        abonos = row[3] or 0
        saldo = row[4] or 0
        
        total_facturado_proveedores += factura
        total_abonado_proveedores += abonos
        saldo_total_proveedores += saldo
        
        if saldo > 0:
            proveedores_con_saldo.append({
                'nombre': row[1],
                'saldo': saldo
            })
    
    # Ordenar proveedores por saldo (de mayor a menor)
    proveedores_con_saldo.sort(key=lambda x: x['saldo'], reverse=True)
    
    # Calcular estadísticas generales de clientes (si existen)
    total_facturado_clientes = 0
    total_abonado_clientes = 0
    saldo_total_clientes = 0
    clientes_con_saldo = []
    
    for row in resumen_clientes:
        if len(row) < 5:
            continue
        factura = row[2] or 0
        abonos = row[3] or 0
        saldo = row[4] or 0
        
        total_facturado_clientes += factura
        total_abonado_clientes += abonos
        saldo_total_clientes += saldo
        
        if saldo > 0:
            clientes_con_saldo.append({
                'nombre': row[1],
                'saldo': saldo
            })
    
    # Ordenar clientes por saldo (de mayor a menor)
    clientes_con_saldo.sort(key=lambda x: x['saldo'], reverse=True)
    
    # Calcular estadísticas de gastos
    total_gastos = 0
    gastos_por_categoria = defaultdict(float)
    gastos_recientes = []
    
    for row in movimientos_gastos:
        if len(row) >= 6:
            fecha, categoria, placa, conductor, precio = row[1], row[2], row[3], row[4], row[5] or 0
            total_gastos += float(precio)
            gastos_por_categoria[categoria] += float(precio)
            
            # Agregar a gastos recientes
            if fecha:
                gastos_recientes.append({
                    'fecha': fecha,
                    'categoria': categoria,
                    'placa': placa,
                    'conductor': conductor,
                    'precio': precio
                })
    
    gastos_categorias = list(gastos_por_categoria.keys())
    gastos_valores = list(gastos_por_categoria.values())

    # Ordenar gastos por fecha (más recientes primero)
    gastos_recientes.sort(key=lambda x: x['fecha'], reverse=True)
    top_gastos = sorted(gastos_recientes, key=lambda x: float(x['precio'] or 0), reverse=True)[:5]
    
    # Calcular porcentajes
    porcentaje_abono_proveedores = (total_abonado_proveedores / total_facturado_proveedores * 100) if total_facturado_proveedores > 0 else 0
    porcentaje_abono_clientes = (total_abonado_clientes / total_facturado_clientes * 100) if total_facturado_clientes > 0 else 0
    
    # Calcular flujo de caja neto (incluyendo gastos)
    flujo_caja_neto = total_abonado_clientes - total_abonado_proveedores - total_gastos
    
    # Obtener movimientos recientes (últimos 10)
    movimientos_recientes = []
    todos_movimientos = []
    
    # Agregar movimientos de proveedores
    for row in movimientos_proveedores:
        if len(row) >= 6:
            fecha, proveedor, detalle, observacion, total = row[1], row[2], row[3], row[4], row[5] or 0
            if fecha:
                todos_movimientos.append({
                    'fecha': fecha,
                    'entidad': proveedor,
                    'tipo': 'Proveedor',
                    'detalle': detalle,
                    'observacion': observacion,
                    'total': total if detalle == 'Factura' else -total
                })
    
    # Agregar movimientos de clientes (si existen)
    for row in movimientos_clientes:
        if len(row) >= 6:
            fecha, cliente, detalle, observacion, total = row[1], row[2], row[3], row[4], row[5] or 0
            if fecha:
                todos_movimientos.append({
                    'fecha': fecha,
                    'entidad': cliente,
                    'tipo': 'Cliente',
                    'detalle': detalle,
                    'observacion': observacion,
                    'total': total if detalle == 'Factura' else -total
                })
    
    # Agregar movimientos de gastos
    for gasto in gastos_recientes:
        todos_movimientos.append({
            'fecha': gasto['fecha'],
            'entidad': gasto['categoria'],
            'tipo': 'Gasto',
            'detalle': 'Gasto',
            'observacion': f"{gasto['placa']} - {gasto['conductor']}" if gasto['placa'] and gasto['conductor'] else gasto['categoria'],
            'total': -float(gasto['precio'] or 0)
        })
    
    # Ordenar movimientos por fecha (más recientes primero)
    todos_movimientos.sort(key=lambda x: x['fecha'], reverse=True)
    movimientos_recientes = todos_movimientos[:10]
      
    seis_meses_atras = datetime.now() - timedelta(days=180)
    facturacion_mensual = defaultdict(float)
    abonos_mensual = defaultdict(float)
    gastos_mensual = defaultdict(float)
    
    for mov in todos_movimientos:
        if mov['fecha'] >= seis_meses_atras:
            mes = mov['fecha'].strftime('%Y-%m')
            if mov['detalle'] == 'Factura':
                facturacion_mensual[mes] += mov['total']
            elif mov['detalle'] == 'Abono':
                abonos_mensual[mes] += abs(mov['total'])
            elif mov['detalle'] == 'Gasto':
                gastos_mensual[mes] += abs(mov['total'])
    
    # Ordenar meses
    meses_ordenados = sorted(facturacion_mensual.keys())
    facturacion_por_mes = [facturacion_mensual[mes] for mes in meses_ordenados]
    abonos_por_mes = [abonos_mensual.get(mes, 0) for mes in meses_ordenados]
    gastos_por_mes = [gastos_mensual.get(mes, 0) for mes in meses_ordenados]
    
    # Calcular métricas adicionales
    ingresos_netos = total_abonado_clientes - total_gastos
    margen_beneficio = (ingresos_netos / total_abonado_clientes * 100) if total_abonado_clientes > 0 else 0
    
    context = {
        # Estadísticas generales
        'total_facturado_proveedores': total_facturado_proveedores,
        'total_abonado_proveedores': total_abonado_proveedores,
        'saldo_total_proveedores': saldo_total_proveedores,
        'porcentaje_abono_proveedores': porcentaje_abono_proveedores,
        
        'total_facturado_clientes': total_facturado_clientes,
        'total_abonado_clientes': total_abonado_clientes,
        'saldo_total_clientes': saldo_total_clientes,
        'porcentaje_abono_clientes': porcentaje_abono_clientes,
        
        'flujo_caja_neto': flujo_caja_neto,
        'total_gastos': total_gastos,
        'ingresos_netos': ingresos_netos,
        'margen_beneficio': margen_beneficio,
        
        # Listas de entidades con saldo
        'proveedores_con_saldo': proveedores_con_saldo[:5],  # Top 5
        'clientes_con_saldo': clientes_con_saldo[:5],  # Top 5
        
        # Datos de gastos
        'gastos_por_categoria': dict(gastos_por_categoria),
        'gastos_categorias': gastos_categorias,  # listo para JS
        'gastos_valores': gastos_valores,
        'top_gastos': top_gastos,
        
        # Movimientos recientes
        'movimientos_recientes': movimientos_recientes,
        
        # Datos para gráficas de tendencia
        'meses_tendencia': meses_ordenados,
        'facturacion_tendencia': facturacion_por_mes,
        'abonos_tendencia': abonos_por_mes,
        'gastos_tendencia': gastos_por_mes,
    }
    
    return render(request, 'index.html', context)

# Vistas específicas para proveedores y clientes
def descargar_excel_proveedor(request):
    return descargar_excel_entidad(request, 'proveedor')

def descargar_excel_cliente(request):
    return descargar_excel_entidad(request, 'cliente')

# Vistas específicas (ahora son simples wrappers de las vistas genéricas)
def MovimientoProveedor(request):
    return movimiento_view(request, 'proveedor')

def MovimientoCliente(request):
    return movimiento_view(request, 'cliente')

def resumen(request):
    return resumen_view(request, 'proveedor')

def resumenCliente(request):
    return resumen_view(request, 'cliente')

def movimientos(request):
    return movimientos_list_view(request, 'proveedor')

def movimientosCliente(request):
    return movimientos_list_view(request, 'cliente')

def agregar_persona(request):
    return agregar_persona_view(request, 'proveedor')

def agregar_persona_Cliente(request):
    return agregar_persona_view(request, 'cliente')

def index(request):
    return index_view(request,'proveedor','cliente','gastos')

def dashboardCliente(request):
    return dashboard_view(request, 'cliente')

def dashboardProveedor(request):
    return dashboard_view(request, 'proveedor')

def guardar_movimiento(request):
    return guardar_movimiento_view(request, 'proveedor')

def guardar_movimiento_cliente(request):
    return guardar_movimiento_view(request, 'cliente')

def editar_movimiento(request, index):
    return editar_movimiento_view(request, 'proveedor', index)

def editar_movimiento_Cliente(request, index):
    return editar_movimiento_view(request, 'cliente', index)

def editar_proveedor(request, id):
    return editar_persona_view(request, 'proveedor', id)

def editar_cliente(request, id):
    return editar_persona_view(request, 'cliente', id)