import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from excelapp.models import Movimiento, Resumen
from django.db import transaction
from django.conf import settings

RUTA_EXCEL = os.path.join(settings.BASE_DIR, 'Financiero.xlsx')

class Command(BaseCommand):
    help = 'Importa movimientos y resumen desde Excel a la base de datos'

    def handle(self, *args, **kwargs):
        if not os.path.exists(RUTA_EXCEL):
            self.stdout.write(self.style.ERROR(f"No existe el archivo Excel en {RUTA_EXCEL}"))
            return

        wb = load_workbook(RUTA_EXCEL)

        movimientos_importados = 0
        resumen_importado = 0

        with transaction.atomic():
            # Importar movimientos
            if 'Proveedores' in wb.sheetnames:
                ws = wb['Proveedores']

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) < 6:
                        continue
                    id_excel, fecha, proveedor, detalle, obs, total = row

                    # Evitar duplicados por ID (en caso de que existan)
                    mov, created = Movimiento.objects.get_or_create(
                        id=id_excel,
                        defaults={
                            'fecha': fecha,
                            'proveedor': proveedor,
                            'detalle': detalle,
                            'obs': obs,
                            'total': total,
                        }
                    )
                    if created:
                        movimientos_importados += 1

            # Importar resumen
            if 'Resumen' in wb.sheetnames:
                ws = wb['Resumen']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) < 4:
                        continue
                    proveedor, facturas, ahorros, saldo = row
                    resumen, created = Resumen.objects.get_or_create(
                        proveedor=proveedor,
                        defaults={
                            'facturas': facturas or 0,
                            'ahorros': ahorros or 0,
                            'saldo': saldo or 0,
                        }
                    )
                    if created:
                        resumen_importado += 1

        self.stdout.write(self.style.SUCCESS(f"Movimientos importados: {movimientos_importados}"))
        self.stdout.write(self.style.SUCCESS(f"Resumen importado: {resumen_importado}"))
