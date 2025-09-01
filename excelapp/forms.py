from django import forms
import os
from openpyxl import load_workbook
from django.conf import settings
from decimal import Decimal

class MovimientoForm(forms.Form):
    DETALLE_CHOICES = (
        ('Ahorro', 'Ahorro'),
        ('Factura', 'Factura'),
    )
    fecha = forms.DateField(
        widget=forms.DateInput(attrs={'type': 'date'}),
        required=True
    )
    proveedor = forms.CharField(max_length=255, label='Proveedor')  # Changed to ChoiceField
    detalle = forms.ChoiceField(choices=DETALLE_CHOICES, label='Detalle')
    obs = forms.CharField(required=False, label='Observación')
    total = forms.DecimalField(
        max_digits=15, decimal_places=0,
        label='Total',
        widget=forms.NumberInput(attrs={
            'class': 'form-control',
            'step': '0.01',
            'min': '0',
            'placeholder': 'Ingrese el total'
        })
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        ruta_excel = os.path.join(settings.BASE_DIR, 'Financiero.xlsx')
        choices = []
        if os.path.exists(ruta_excel):
            wb = load_workbook(ruta_excel)
            if 'Resumen' in wb.sheetnames:
                ws = wb['Resumen']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre = row[0]
                    if nombre:
                        choices.append((nombre, nombre))
        self.fields['proveedor'].choices = choices
        # Estilos Bootstrap
        self.fields['proveedor'].widget.attrs.update({'class': 'form-control'})
        self.fields['detalle'].widget.attrs.update({'class': 'form-select'})
        self.fields['obs'].widget.attrs.update({'class': 'form-control'})

class MovimientoClienteForm(forms.Form):
    DETALLE_CHOICES = (
        ('Ahorro', 'Ahorro'),
        ('Factura', 'Factura'),
    )
    fecha = forms.DateField(
        widget=forms.DateInput(attrs={'type': 'date'}),
        required=True
    )
    proveedor = forms.CharField(max_length=255, label='Proveedor')  # Changed to ChoiceField
    detalle = forms.ChoiceField(choices=DETALLE_CHOICES, label='Detalle')
    obs = forms.CharField(required=False, label='Observación')
    total = forms.DecimalField(
        max_digits=15, decimal_places=0,
        label='Total',
        widget=forms.NumberInput(attrs={
            'class': 'form-control',
            'step': '0.01',
            'min': '0',
            'placeholder': 'Ingrese el total'
        })
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        ruta_excel = os.path.join(settings.BASE_DIR, 'Financiero.xlsx')
        choices = []
        if os.path.exists(ruta_excel):
            wb = load_workbook(ruta_excel)
            if 'ResumenCliente' in wb.sheetnames:
                ws = wb['ResumenCliente']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre = row[0]
                    if nombre:
                        choices.append((nombre, nombre))
        self.fields['proveedor'].choices = choices
        # Estilos Bootstrap
        self.fields['proveedor'].widget.attrs.update({'class': 'form-control'})
        self.fields['detalle'].widget.attrs.update({'class': 'form-select'})
        self.fields['obs'].widget.attrs.update({'class': 'form-control'})

def clean_total(self):
    total_raw = self.cleaned_data['total']
    if isinstance(total_raw, (int, float, Decimal)):
        return total_raw

    from decimal import Decimal, InvalidOperation
    import re

    # Quitar símbolo de pesos, puntos y espacios
    total_str = re.sub(r'[^\d]', '', str(total_raw))  # solo deja números

    if not total_str:
        raise forms.ValidationError("Introduzca un número válido.")

    try:
        return Decimal(total_str)
    except InvalidOperation:
        raise forms.ValidationError("Introduzca un número válido.")

class ProveedorForm(forms.Form):
    nombre = forms.CharField(
        label='Nombre del Proveedor',
        max_length=100,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Nombre del proveedor'})
    )